#!/usr/bin/python
# coding: utf8

import sys, os
import re
import getopt
import json
from openpyxl import load_workbook

from class_mapping import OWLResourceManager, OWLClass, OWLIndividual
from label_mapping import LABEL_MAPPING_PREFIX, \
                          LABEL_MAPPING_SUFFIX, \
                          LABEL_MAPPING_WORD, \
                          LABEL_MAPPING_REPLACE
from subclass_mapping import SHOP_CLASS_REPLACEMENT, \
                             SHOP_SUBCLASS_MAPPING

def clean_label(label):
  words = label.replace('-',' '). \
                replace('.',' '). \
                replace('(',' '). \
                replace(')',' '). \
                replace('+',' oder '). \
                replace('&',' oder '). \
                replace('/',' oder '). \
                split()
  # replace prefix (first word)
  try:    words[0]  = LABEL_MAPPING_PREFIX[words[0]]
  except: pass
  # replace suffix (last word)
  try:    words[-1] = LABEL_MAPPING_SUFFIX[words[-1]]
  except: pass
  # replace words
  for i in range(len(words)):
    try:    words[i] = LABEL_MAPPING_WORD[words[i]]
    except: pass
  # Avoid that words are repeated and
  # that the next is the suffix of the previous one
  last = ''
  withoutRepetition = []
  for x in words:
    if last.lower().endswith(x.lower()): continue
    last=x
    withoutRepetition.append(x)
  clean = ' '.join(withoutRepetition).strip()
  return clean
  #try:    return LABEL_MAPPING_REPLACE[clean]
  #except: return clean

def label_replace(column,label):
  try:
    return LABEL_MAPPING_REPLACE[label]
  except:
    return label

class ProductTableColumn:
  """
  bla bla bla
  """
  def __init__(self, name, labels,
               coltype='class',
               parents=[],
               role=None,
               param=None, paramType=None,
               cls=None):
    self.name      = name
    self.labels    = labels
    self.coltype   = coltype
    self.parents   = parents
    self.role      = role
    self.cls       = cls
    self.param     = param
    self.paramType = paramType
  
  def format_label(self, label):
    return clean_label(label)
  
  def match_label(self, label):
    for x in self.labels:
      if str(label)==str(x): return True
    return False
  
  def __repr__(self):
    return self.name

class ProductTable:
  """
  bla bla bla
  """
  COLUMN_META_DATA = [
    ProductTableColumn('assortment',
      labels=['Warensortiment'],
      cls='Product'),
    ProductTableColumn('area',
      labels=['Warenbereich'],
      cls='Product'
    ),
    ProductTableColumn('subArea',
      labels=['Warenunterbereich'],
      cls='Product',
      parents=['area'],
      role='type'),
    ProductTableColumn('generalType',
      labels=['Warenklasse'],
      cls='Product'
    ),
    # TODO: also include information from specificType column
    #ProductTableColumn('specificType',
    #  labels=['Warengruppe'],
    #  parents=['generalType'],
    #  role='type'),
    #ProductTableColumn('brand',
    #  labels=['Marke Dachmarke'],
    #  coltype='instance',
    #  cls='ProductBrand',
    #  role='brand'),
    #ProductTableColumn('subBrand',
    #  labels=['Marke Submarke'],
    #  coltype='instance',
    #  cls='ProductBrand',
    #  role='brand'),
    ProductTableColumn('width',
      labels=['Artikel Breite (DANf) cm'],
      coltype='parameter',
      param='widthOfProduct',
      paramType='&xsd;float'),
    ProductTableColumn('height',
      labels=['Artikel Hoehe (DANf) cm'],
      coltype='parameter',
      param='heightOfProduct',
      paramType='&xsd;float'),
    ProductTableColumn('depth',
      labels=['Artikel Laenge (DANf) cm'],
      coltype='parameter',
      param='depthOfProduct',
      paramType='&xsd;float'),
    ProductTableColumn('article',
      labels=['Artikel'],
      coltype='annotation',
      role='label'),
    ProductTableColumn('articleNumber',['DAN', 'EAN'],coltype='instance'),
  ]
  
  
  def __init__(self, resourceManager, whitelist, meshPrefix, xlsxFile, xlsxTable, lang='de'):
    self.resourceManager = resourceManager
    self.table = load_workbook(xlsxFile)[xlsxTable]
    self.lang  = lang
    self.rawLabels = set()
    self.class_articles = {}
    self.whitelist = whitelist
    self.meshPrefix = meshPrefix
    # create ordered list of columns we have some meta information about
    self.header = []
    for row in self.table.iter_rows(min_row=1, max_row=1):
      self.header=map(lambda x: self.find_column(x.value), row)
  
  def article(self, articleNumber):
    articleClass = "ProductWithAN"+articleNumber
    try: return self.resourceManager.owlClasses[articleClass]
    except:
      article = OWLClass(articleClass, self.resourceManager.ontologyPrefix)
      article.has_type('Product')
      article.has_object_value("articleNumberOfProduct", self.article_number(articleNumber))
      self.resourceManager.owlClasses[articleClass] = article
      return article
  
  def article_number(self, articleNumber):
    x = self.resourceManager.get_instance('ArticleNumber',
                                          'ArticleNumber_'+articleNumber)
    x.has_data_value('articleNumberString',
                     'http://knowrob.org/kb/shop.owl#dan',
                     articleNumber)
    return x

  def unique_names(self, column):
    """ read all values of a column (may be many) """
    names=set()
    for row in self.table.iter_rows(min_row=2):
      names.add(self.read_value(row, column))
    out = list(names)
    out.sort()
    return out
  
  def read(self):
    """ read in the complete table, create ProductClass instances on the way """
    for row in self.table.iter_rows(min_row=2):
      # first get article class
      articleNumber = str(self.read_value(row, 'articleNumber')).zfill(6)
      # skip non whitelisted
      if self.whitelist != None and not articleNumber in self.whitelist: continue
      articleClass  = self.article(articleNumber)
      # check if we have a mesh value in the whitelist
      if self.whitelist != None:
        meshPath = self.whitelist[articleNumber]
        if len(meshPath)>4: articleClass.meshPath = self.meshPrefix + meshPath
      for i in range(len(self.header)):
        column = self.header[i]
        if column!=None and (column.role!=None or column.param!=None):
          self.read_cell(articleClass,row,column)

  def read_cell(self,article,row,column):
    rawLabel = self.read_value(row,column)
    label = column.format_label(rawLabel)
    if column.coltype=='class':
      label = label_replace(column,label)
      cellClass = self.resourceManager.get_translated(label, lang=self.lang)
      if cellClass==None: return None
      cellClass.has_type('Product')
      self.add_class_article(article,cellClass,row,column)
      #self.read_cell_types(article,cellClass,row,column)
      self.read_cell_property(article,cellClass,column)
      self.rawLabels.add(rawLabel)
      # columns may have parent classes in the upper shopping ontology
      if column.cls!=None: cellClass.has_type(column.cls)
      return cellClass.name
    elif column.coltype=='instance' and column.cls!=None:
      cellInstance = self.resourceManager.get_instance(label, column.cls)
      if cellInstance==None: return None
      self.read_cell_types(article,cellInstance,row,column)
      self.read_cell_property(article,cellInstance,column)
    elif column.coltype=='parameter':
      # HACK cm -> m
      if column.paramType=="&xsd;float":
        num = float(rawLabel)/100.0
        rawLabel = str(num)
      self.read_cell_property(article,rawLabel,column)
    return label

  def add_class_article(self,article,cellClass,row,column):
    try:
      l = self.class_articles[cellClass.name]
    except:
      l = list()
      self.class_articles[cellClass.name] = l
    l.append((article,cellClass,row,column))

  def read_cell_types(self,article,resource,row,column):
    for x in column.parents:
      clsName = self.read_cell(article,row,self.get_column(x))
      if clsName==resource.name or clsName==None: continue
      if self.resourceManager.subclass_of(clsName, resource.name):
        print("WARN: mutual subclassOf relation between " + str((clsName, resource.name)))
      else:
        # HACK a is subclass b if a's name contains b's name
        if resource.name in clsName:
          cls = self.resourceManager.owlClasses[clsName]
          cls.has_type(resource.name)
        else:
          resource.has_type(clsName)

  def read_cell_property(self,article,resource,column):
    if column.role=='type':
      article.has_type(str(resource))
    elif column.role!=None:
      article.has_object_value(column.role, str(resource))
    elif column.param!=None and column.paramType!=None:
      article.has_data_value(column.param, column.paramType, str(resource))

  def read_value(self, row, column):
    index=self.get_column_index(column)
    if index==None: return None
    val=row[index].value
    if not isinstance(val, basestring): val=unicode(val)
    return val.encode('utf-8')
  
  def get_column_index(self,col):
    i=0
    for x in self.header:
      if str(col)==str(x): return i
      i+=1
  
  def get_column(self,col):
    for x in self.header:
      if str(col)==str(x): return x
  
  def find_column(self,x):
    for col in ProductTable.COLUMN_META_DATA:
      if col.match_label(x): return col


def main(argv):
  resourceManager = OWLResourceManager(ontologyPrefix='shop', \
                                       ontologyURI='http://knowrob.org/kb/shop.owl',\
                                       baseClass='Product')
  resourceManager.add_import('package://knowrob_refills/owl/shop.owl')
  resourceManager.add_import('package://knowrob_refills/owl/dm-market.owl')
  resourceManager.add_class_mapping(SHOP_CLASS_REPLACEMENT)
  resourceManager.add_subclass_mapping(SHOP_SUBCLASS_MAPPING)
  
  outDir = os.path.abspath(os.path.dirname(os.path.realpath(__file__)) + "/../owl/")
  # handle commandline parameters
  try:
    opts, args = getopt.getopt(argv, "tcsw:p:l:i:o:m:",
      ["taxonomy", "catalog", "subclasses", "whitelist=", "mesh-prefix=", "list=", "input=", "output-dir=", "output-mode="])
  except getopt.GetoptError:
    sys.exit(2)
  opt_taxonomy = False
  opt_catalog = False
  opt_subclasses = False
  opt_mesh_prefix = "package://refills_models/"
  opt_whitelist = {}
  opt_list = []
  opt_tables = []
  for opt, arg in opts:
    if opt in ("-t", "--taxonomy"):
      opt_taxonomy = True
    elif opt in ("-s", "--subclasses"):
      opt_subclasses = True
    elif opt in ("-c", "--catalog"):
      opt_catalog = True
    elif opt in ("-w", "--whitelist"):
      opt_whitelist = json.load(open(arg))
    elif opt in ("-l", "--list"):
      opt_list.append(arg)
    elif opt in ("-i", "--input"):
      opt_tables.append(arg)
    elif opt in ("-o", "--output-dir"):
      outDir = arg
    elif opt in ("-m", "--output-mode"):
      resourceManager.set_output_mode(arg)
  # instantiate tables
  tables = []
  for table_arg in opt_tables:
    x = table_arg.split(":")
    tables.append(ProductTable(resourceManager, opt_whitelist, opt_mesh_prefix, x[0], x[1]))
  # print out unique names
  for l in opt_list:
      for t in tables:
        column = t.get_column(l)
        if column==None: continue
        for x in t.unique_names(l):
          x0=str(x)
          clean=clean_label(x0)
          if clean=="": continue
          label = column.format_label(x0)
          if not label in LABEL_MAPPING_REPLACE:
            x1=str(resourceManager.get_name(label))
            print("    '"+clean+"': ('"+x1+"',[]),")
      resourceManager.translator.save()
  if opt_subclasses:
      print("Finding classes without defined sublcass...")
      clsSet=set()
      for t in tables:
        t.read()
        for articleList in t.class_articles.values():
          for (article,cls,row,col) in articleList:
            if cls.name.startswith('ProductWithAN'): continue
            if cls.name in SHOP_SUBCLASS_MAPPING: continue
            t.read_cell_types(article,cls,row,col)
            clsSet.add(cls.name)
        resourceManager.translator.save()
      resourceManager.cleanup_subclasses()
      clsSet = list(clsSet)
      clsSet.sort()
      for clsName in clsSet:
        cls = resourceManager.owlClasses[clsName]
        print("    '"+cls.name+"': "+str(list(cls.types))+",")
      print("Found " + str(len(clsSet)) + " classes with unspecified parent classes.")
  # separate product catalog from taxonomy
  isProductClass  = lambda e: e.name.startswith('ProductWithAN')
  isIndividual    = lambda e: isinstance(e, OWLIndividual)
  filter1         = lambda e: isIndividual(e) or isProductClass(e)
  isClass         = lambda e: isinstance(e, OWLClass)
  filter2         = lambda e: isClass(e) and not isProductClass(e)
  # print out unique names
  if opt_taxonomy or opt_catalog:
      print("Reading tables (this may take a while)...")
      for t in tables:
        t.read()
        resourceManager.translator.save()
      resourceManager.cleanup_subclasses()
      print("Dumping ontology to " + outDir)
  if opt_taxonomy: resourceManager.dump(outDir+"/product-taxonomy.owl", \
                                        'http://knowrob.org/kb/product-taxonomy.owl', filter1)
  if opt_catalog:  resourceManager.dump(outDir+"/product-catalog.owl", \
                                        'http://knowrob.org/kb/product-catalog.owl', filter2)
  if opt_taxonomy or opt_catalog:
      voidFilter = lambda e: False
      # some debug output at the end
      print("Ontology statistics:")
      print("    total classes:     " + str(resourceManager.class_count(voidFilter)))
      print("    total individuals: " + str(resourceManager.individual_count(voidFilter)))
      print("    product classes:   " + str(resourceManager.entity_count(lambda e: not isProductClass(e))))
      uniqueNames = set()
      for t in tables:
         for x in t.rawLabels: uniqueNames.add(x)
      print("    raw taxonomy:      " + str(len(uniqueNames)))
      print("    cleaned taxonomy:  " + str(resourceManager.entity_count(filter1)))
      print("    catalog resources: " + str(resourceManager.entity_count(filter2)))

if __name__ == "__main__":
  main(sys.argv[1:])

